#!/usr/bin/env python3
import sys

try:
    import openpyxl
except ImportError:
    print("A biblioteca 'openpyxl' é necessária para a utilização deste programa. Verifique se a mesma está instalada em seu sistema")
#Biblioteca para manipular .xlsx

try:
    import numpy as np
    from scipy.stats import pearsonr
except ImportError:
    print("A biblioteca 'scipy' é necessária para a utilização deste programa. Caso queira instalá-lo, é necessário utilizar o comando 'pip install scipy' em seu prompt de comando")
#Biblioteca para o calculo de correlação de Pearson

try:
    import seaborn as sns
    import matplotlib.pyplot as plt
    import pandas as pd
except ImportError:
    print("Pacote seaborn ausente")
#Biblioteca para plotagem de um heatmap com os resultados do teste de correlação de Pearson

planilha = sys.argv[1]

def mean_grouped(dados):
    mean_values = {}
    valores_tempo = {}
    for linha in range(2, dados.max_row + 1):
        valor_tempo = dados.cell(row = linha, column = 2).value
        valor = dados.cell(row = linha, column = 3).value
        if valor_tempo is not None and valor is not None:
#Criando/adicionando valores ao dicionário valores_tempo[] pela verificação se o dicionário contém a chave ou não
            valor = float(valor)
            if valor_tempo in (0, 1, 2, 4, 8):
                if valor_tempo not in valores_tempo:
                    valores_tempo[valor_tempo] = [valor]
                else:
                    valores_tempo[valor_tempo].append(valor)
        else:
            continue
#Calculando a média entre 3 valores 
    for tempo, valores in valores_tempo.items():
        medias_tempo = []
        for i in range(0, len(valores), 3):
            media = sum(valores[i:i+3]) / 3
            medias_tempo.append(media)
        mean_values[tempo] = medias_tempo
    return(mean_values)

#Ínicio do script
try:
    matrizes = []
    nomes_paginas = ['FvFm', 'ChlIdx', 'AriIdx']
#lista das matrizes geradas para comparação
    matriz_correlacao = np.zeros((len(nomes_paginas), len(nomes_paginas)))
#matriz de correlação para plotagem do heatmap - np.zeros cria uma matriz preenchida por zeros
    arquivo = openpyxl.load_workbook(planilha)
    print(f"O arquivo foi aberto e salvo com sucesso!\n")
    pages_names = arquivo.sheetnames
    for name in pages_names:
        print(f"Processando a página {name}")
        dados = arquivo[name]
        resultado = mean_grouped(dados)
        matriz = [[0.0 for _ in range(5)] for _ in range(5)]
#criando matrizes separadamente para adicionar à lista 'matrizes'

#loop para adicionar os valores do output da função em uma matriz 5x5
        for i, tempo in enumerate(sorted(resultado.keys())):
            medias = resultado[tempo]
            matriz[i] = medias
#adicionando as matrizes à lista 'matrizes' criada anteriormente
        matrizes.append(matriz)
#loop para printar as linhas da matriz alimentada com os valores
        for linha in matriz:
            row_edit = [round(media, 3) for media in linha]
            print(row_edit)
        print("\n")
    print(matrizes)
    print("\n")

    for i in range(len(nomes_paginas)):
        for j in range(i+1, len(nomes_paginas)):
            nome1, nome2 = nomes_paginas[i], nomes_paginas[j]
            correlacao = pearsonr(np.array(matrizes[i]).flatten(), np.array(matrizes[j]).flatten())
            print(f"Correlação entre {nome1}, {nome2}: {correlacao}")
            matriz_correlacao[i, j] = matriz_correlacao[j, i] = correlacao[0]
#comparar nome1 com nome2 é o equivalente a comparar nome2 com nome1

    correlacao_df = pd.DataFrame(matriz_correlacao, index = nomes_paginas, columns = nomes_paginas)

    plt.figure(figsize=(10, 8))
    sns.heatmap(correlacao_df, annot=True, cmap='coolwarm', fmt=".2f", linewidths=.5)
    plt.title('Matriz de Correlação entre as Páginas')
    plt.savefig('correlation_heatmap.png')
    

except IOError as e:
    print(f"Erro ao ler o arquivo: {e}")
except Exception as e:
    print(f"Erro desconhecido... {e}")
